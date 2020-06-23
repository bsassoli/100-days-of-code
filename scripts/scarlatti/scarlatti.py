"""
A simple app to manage todos for our new home

"""
import os
from openpyxl import Workbook, load_workbook


class Room():

    """Summary

    Attributes:
        name (TYPE): Description
    """

    def __init__(self, name):
        """Summary

        Args:
            name (TYPE): Description
        """
        self.name = name


def file_creation():
    """
    user decides whether to work with an existing file or to create a new one
    returns a workbook

    Returns:
        TYPE: Description
    """
    inp = ''
    while inp not in ['Yes', 'No']:
        inp = input('Would you like to start a new project? (Yes/No)? ')
        if inp == 'Yes':
            path = str(input('What do you want you filename to be? '))
            path += '.xlsx'
            workbook = Workbook()
            workbook.save(filename=path)

        elif inp == 'No':
            files = os.listdir(path='.')
            for n in range(len(files)):
                print(str(n) + ' ' + files[n])
            num = int(input(
                'Please choose the file you would like to work with'))
            path = files[num]
    print(path)
    return path


def choose_action():
    """
    """

    print('What would you like to do?')
    print('Add a room: 1')
    print('Add an action: 2')
    print('Modify an action: 3')
    print('View a room: 4')
    print('View an action: 5')
    user_choice=int(input('Please choose an option:'))
    choice_dict = {1: 'add_room', 2: 'add_action', 3 : 'modify_action', 
                   4: 'view_room', 5: 'view_action'}
    return choice_dict[user_choice]



def create_rooms(path):
    """
    creates sheets for each room based on user input.
    Returns a dictionary of ```Room``` objects

    Args:
        path (string): the relative path of the file to be opened

    Returns:
        Dictionary: a dictionary of room objects
    """
    rooms = []
    answer = ''
    workbook = load_workbook(filename=path)
    while answer != 'Done':
        answer = input('Write the name of a room (or "Done" when done): ')
        if answer != 'Done':
            rooms.append(answer)

    for room in rooms:
        workbook.create_sheet(room)

    rooms_dic = {room: Room(room) for room in rooms}

    workbook.save(filename=path)
    return rooms_dic


choose_action()
